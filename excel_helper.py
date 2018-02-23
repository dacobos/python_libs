import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from copy import copy



def read_data(filename):
    data = []
    # Read an XLS or XLSX File and create a List of the file
    book = xlrd.open_workbook(filename)
    xl_sheet = book.sheet_by_index(0)
    num_rows = xl_sheet.nrows
    for i in range(0, num_rows):
        data.append(xl_sheet.row_values(i))
    print "excel_helper.py - result: Data read from: " + str(filename)
    return data

def getColumnId(data, columnName):
    # Based on a key get the row number and get that row values
    for line in data:
        for i in range(len(line)):
            try:
                if columnName in line[i]:
                    return i
            except TypeError:
                continue
    return None

def getRowsByKey(data, key):
    # Based on a key search for row numbers that match the key
    Rows = {}
    rowNumbers = []
    for i in range(len(data)):
        row = data[i]
        try:
            if any(key in l for l in row):
                rowNumbers.append(i)

        except TypeError as e:
            # print e
            pass

    Rows[key] = rowNumbers
    return Rows

def getRowsByColumn(data, columnId, value):
    # Search on every row on an specific column for certain value and return RowNumber
    rows = []
    for r in range(len(data)):
        row = data[r]
        if row[columnId] == value:
            rows.append(r)
    return rows

def getValues(data, rows):
    # Based on a row number get the values of that list
    values = []
    for key in rows:
        rowNumbers = rows[key]
        for row in rowNumbers:
            values.append(data[row])
    return values

def getCellValue(data, rowId, columnId):
    # Based on a row number and column headerId get the cell value
    cellValue = data[rowId][columnId]
    return cellValue

def updateCell(data, rowId, columnId, newValue):
    # Based on a row number and a column headerId update the cell value
    try:
        data[rowId][columnId] = newValue
        print "excel_helper.py - result: Cell updated with: " + str(newValue)
        return True
    except:
        print "excel_helper.py - result: Could't update cell with: " + str(newValue)
        return False

def readXlsxSheet(filename):
    workbook = load_workbook(filename)
    worksheet = workbook.active
    print "excel_helper.py - result: Worksheet loaded " + str(worksheet)
    return worksheet

def writeXlsx(data, newfilename):
    # newfilename = filename.split('.')[0]+'_codigos_sap.xlsx'
    wb = Workbook()
    ws = wb.active
    for i in range(len(data)):
        ws.append(data[i])
    wb.save(newfilename)
    print "excel_helper.py - result: Workbook Saved " + str(newfilename)
    return newfilename

def duplicateXlsx(template_file, new_data_file):
    # Duplicate xlsx file with cell borders, fonts and merged cells using new_data_file
    # Read template
    wb_template = load_workbook(template_file)
    ws_template = wb_template.active

    # ws = wb_template.worksheets[0]
    # print ws.merged_cell_ranges

    # Create new WB
    wb_new = Workbook()
    ws_new = wb_new.active

    # Copy all the merged_cells to the new document
    merged = ws_template.merged_cell_ranges
    for elem in merged:
        ws_new.merge_cells(elem)



    row_idx = 0
    for row in ws_template.rows:
        data_row = new_data_file[row_idx]
        row_idx += 1
        # change the row height
        ws_new.row_dimensions[1].height = ws_template.row_dimensions[1].height
        for c in range(len(row)):
            cell = row[c]
            try:
                new_cell = ws_new.cell(row=row_idx, column=cell.col_idx, value=data_row[c])
            except:
                new_cell = ws_new.cell(row=row_idx, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for col in ws_template.columns:
        col_letter = str(get_column_letter(col[0].col_idx))
        ws_new.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width

    # Copy the borders of merged cell
    for elem in merged:
        cell_range = ws_new[elem]
        bordered = ws_new[elem.split(':')[0]]
        for i in range(len(cell_range)):
            cell = cell_range[i]
            for c in cell:
                c.border = copy(bordered.border)
        print "excel_helper.py - result: Workbook Created from: " + str(template_file)
        return wb_new

def saveWb(wb, newfilename):
    ws = wb.active
    wb.save(newfilename)
    print "excel_helper.py - result: Workbook Saved to:" + str(newfilename)
    return

def dataFromDic(dictionary,sheet_by_sub_key):
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)
    command_list = []
    row_idx = 1
    col_idx = 0
    for key in dictionary:
        if sheet_by_sub_key:
            # Using sub key as sheet names
            for sub_key in dictionary[key]:
                if any(sub_key in l for l in command_list):
                    continue
                else:
                    command_list.append(sub_key)
                    ws_name = sub_key
                    ws = wb.create_sheet(ws_name)
        else:
            # Using external key as sheet names
            ws_name = key
            ws = wb.create_sheet(ws_name)



    if sheet_by_sub_key:
        sheetnames = wb.sheetnames
        for sheet in sheetnames:
            col_idz = 2
            row_idx = 2
            ws=wb[sheet]
            for key in dictionary:
                try:
                    data_list = dictionary[key][sheet]
                    # write the first key to ws
                    ws.cell(row=row_idx, column=col_idz, value=key)
                    ws.cell(row=row_idx, column=col_idz+1, value=sheet)
                    col_idx = col_idz+1
                    for i in range(len(data_list)):
                        col_idx += 1
                        # Write each elem to ws
                        new_value = data_list[i]
                        if type(new_value) == type([]):
                            col_idy = col_idz+1
                            # Write separating columns
                            for elem in new_value:
                                col_idy += 1
                                ws.cell(row=row_idx, column=col_idy, value=elem)
                            row_idx += 1
                        else:
                            # Write separating rows
                            ws.cell(row=row_idx, column=col_idx, value=new_value)
                except KeyError:
                    pass
                row_idx += 1
    else:
        sheetnames = wb.sheetnames
        for sheet in sheetnames:
            col_idz = 2
            row_idx = 2
            ws=wb[sheet]
            for sub_key in dictionary[sheet]:
                try:
                    data_list = dictionary[sheet][sub_key]
                    # write the first key to ws
                    ws.cell(row=row_idx, column=col_idz, value=sheet)
                    ws.cell(row=row_idx, column=col_idz+1, value=sub_key)
                    col_idx = col_idz+1
                    for i in range(len(data_list)):
                        col_idx += 1
                        # Write each elem to ws
                        new_value = data_list[i]
                        if type(new_value) == type([]):
                            # Result with double level list
                            col_idy = col_idz+1
                            # Write separating columns
                            for elem in new_value:
                                col_idy += 1
                                ws.cell(row=row_idx, column=col_idy, value=elem)
                            row_idx += 1
                        else:
                            # Result with single level list
                            ws.cell(row=row_idx, column=col_idx, value=new_value)
                except KeyError:
                    pass

                row_idx += 1
    print "excel_helper.py - result: Workbook Created " + str(wb)
    return wb
